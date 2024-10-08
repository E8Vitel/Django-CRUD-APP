import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from django.conf import settings
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import send_mail
from openpyxl.styles import NamedStyle
from bodega.decorators import admin_required
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import get_object_or_404, redirect, render
from .forms import CustomAuthenticationForm, ProductoForm, SolicitudForm, UnidadForm
from .models import DetalleSolicitud, Producto, Categoria, Historial, DetallesHistorial, Solicitud, Unidad, UserProfile

# Create your views here.

# Método para mostrar el formulario de inicio de sesión
def view_login(request):
    # Si el método es POST, intenta autenticar al usuario
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('view_productos')
    else:
        form = CustomAuthenticationForm()

    return render(request, 'login.html', {'form': form})

# Método para cerrar sesión del usuario
def view_logout(request):
    logout(request)
    return redirect('login')

# Vista para mostrar el historial de transacciones (requiere permisos de administrador)
@admin_required
def view_historial(request):
    historiales = Historial.objects.all()
    context = {'historiales': historiales}
    return render(request, 'historial.html', context)

# Vista para mostrar los productos disponibles
@login_required
def view_productos(request):
    categorias = Categoria.objects.all()
    categoria_filter = request.GET.get('categoria', 'todo')
    productos = Producto.objects.all()
    query = request.GET.get('q', '')

    # Filtrar productos según la categoría y la búsqueda
    if categoria_filter != 'todo':
        productos = Producto.objects.filter(categoria_id=categoria_filter, nombre_producto__icontains=query)
    else:
        productos = Producto.objects.filter(nombre_producto__icontains=query)

    # Calcular el precio total de cada producto
    for producto in productos:
        producto.precio_total = producto.cantidad * producto.precio

    context = {'categorias': categorias, 'productos': productos, 'categoria_filter': categoria_filter, 'producto_modal': None}

    return render(request, 'bodega.html', context)

# Vista para crear nuevos productos (requiere permisos de administrador)
@admin_required
def view_create_productos(request):
    productos = Producto.objects.all()
    categorias = Categoria.objects.all()
    unidades = Unidad.objects.all()
    return render(request, 'productos.html', {'productos': productos, 'categorias': categorias, 'unidades': unidades})

# Vista para crear productos existentes (requiere permisos de administrador)
@admin_required
def view_create_existing_productos(request):
    productos = Producto.objects.all()
    categorias = Categoria.objects.all()
    unidades = Unidad.objects.all()
    return render(request, 'productosEx.html', {'productos': productos, 'categorias': categorias, 'unidades': unidades})

# Vista para mostrar la salida de productos y permitir su registro (requiere inicio de sesión)
@login_required
def view_productos_output(request):
    productos = Producto.objects.all()
    categorias = Categoria.objects.all()
    form = SolicitudForm()

    return render(request, 'salidaProd.html', {'productos': productos, 'categorias': categorias,'form': form})

# Vista para mostrar las unidades disponibles (requiere inicio de sesión)
@login_required
def view_units(request):
    unidades = Unidad.objects.all()
    perfiles = UserProfile.objects.all()
    return render(request, 'unidades.html', {'unidades': unidades, 'perfiles': perfiles})

# Vista para mostrar las solicitudes pendientes (requiere permisos de administrador)
@admin_required
def view_solicitudes(request):
    if not request.user.is_staff:
        return redirect('view_productos')

    solicitudes_pendientes = Solicitud.objects.filter(estado='pendiente')
    
    return render(request, 'solicitudes.html', {'solicitudes_pendientes': solicitudes_pendientes})

# Método para crear un nuevo producto (requiere permisos de administrador)
@admin_required
def create_producto(request):
    if request.method == 'POST':
        # Obtener los datos del formulario
        nombre = request.POST.getlist('nombreProducto')
        descripcion = request.POST.getlist('descripcion')
        precio = request.POST.getlist('precio')
        cantidad = request.POST.getlist('cantidad')
        categoria_id = request.POST.getlist('categoria')
        proveedores_ids = request.POST.getlist('proveedores')

        # Crear un nuevo historial de transacciones
        historial = Historial.objects.create(fecha=datetime.now())

        # Verificar si ya existe un producto con el mismo nombre
        if Producto.objects.filter(nombre_producto=nombre).exists():
            raise ValidationError("Ya existe un producto con este nombre.")
        
        # Crear nuevos productos con sus respectivos detalles
        for nombre, descripcion, precio, cantidad, categoria_id in zip(nombre, descripcion, precio, cantidad, categoria_id):
            categoria = Categoria.objects.get(id=categoria_id)
            producto = Producto.objects.create(
                nombre_producto=nombre,
                descripcion=descripcion,
                precio=float(precio),
                cantidad=int(cantidad),
                categoria=categoria,
            )

            producto.proveedores.set(proveedores_ids)

            DetallesHistorial.objects.create(
                historial=historial,
                producto=producto,
                nombre_producto=producto.nombre_producto,
                cantidad=producto.cantidad,
                precio_unitario=producto.precio,
                unidad=producto.proveedores.first()
            )

        return redirect('view_productos')

    return render(request, 'bodega.html')

# Método para crear productos existentes (requiere permisos de administrador)
@admin_required
def create_existing_productos(request):
    productos = Producto.objects.all()

    if request.method == 'POST':
        productos_data = request.POST.getlist('productos[]')
        precios = request.POST.getlist('precios[]')
        cantidad_cajas = request.POST.getlist('cajas[]')
        cantidades = request.POST.getlist('cantidades[]')
        proveedores_id = request.POST.getlist('proveedores[]')

        # Crear un nuevo historial de transacciones
        historial = Historial.objects.create(fecha=datetime.now())

        # Iterar sobre los datos recibidos para crear los detalles del historial y actualizar los productos existentes
        for producto_id, nuevo_precio, cajas, cantidad, proveedor in zip(productos_data, precios, cantidad_cajas, cantidades, proveedores_id):
            producto_existente = Producto.objects.get(id=producto_id)
            unidad = Unidad.objects.get(id=proveedor)
            if cajas:
                cantidad = int(cantidad) * int(cajas)
            else:
                cantidad = int(cantidad)

            detalle_historial = DetallesHistorial(
                historial=historial,
                producto=producto_existente,
                nombre_producto=producto_existente.nombre_producto,
                cantidad=int(cantidad),
                unidad=unidad
            )
            if nuevo_precio:
                try:
                    detalle_historial.precio_unitario = float(nuevo_precio)
                    producto_existente.precio = float(nuevo_precio)
                except ValueError:
                    detalle_historial.precio_unitario = producto_existente.precio
            else:
                detalle_historial.precio_unitario = producto_existente.precio

            detalle_historial.save()
            producto_existente.cantidad += int(cantidad)
            producto_existente.save()

        return redirect('view_productos')

    return render(request, 'bodega.html', {'productos': productos})

# Método para procesar la salida de productos (requiere inicio de sesión)
@login_required   
def producto_output(request):
    if request.method == 'POST':
        usuario = request.user
        solicitud = Solicitud.objects.create(fecha=datetime.now(), usuario=usuario, estado='pendiente')

        productos_seleccionados = request.POST.getlist('productos')
        cantidades = request.POST.getlist('cantidades')
        cajas = request.POST.getlist('cajas')

        # Iterar sobre los productos seleccionados y crear detalles de solicitud asociados
        for producto_id, cajas, cantidad in zip(productos_seleccionados, cajas, cantidades):
            if cajas:
                cantidad = int(cantidad) * int(cajas)
            else:
                cantidad = int(cantidad)
            if cantidad > 0:
                producto = Producto.objects.get(pk=producto_id)
                DetalleSolicitud.objects.create(
                    solicitud=solicitud,
                    producto=producto,
                    cantidad=cantidad
                )

        # Enviar correo electrónico de notificación
        enviar_correo(solicitud.detalles_solicitud.all())

        messages.success(request, "Solicitud enviada correctamente.")
        return redirect('view_productos')

    productos = Producto.objects.all()
    return render(request, 'salidaProd.html', {'productos': productos})

# Método para crear una nueva unidad (requiere permisos de administrador)
@admin_required
def create_unit(request):
    if request.method == 'POST':
        nombre_unidad = request.POST.get('nombre_unidad')

        # Verificar si ya existe una unidad con el mismo nombre
        if Unidad.objects.filter(nombre_Unidad=nombre_unidad).exists():
            raise ValidationError("Ya existe una unidad con este nombre.")
        
        # Crear una nueva unidad
        Unidad.objects.create(
            nombre_Unidad=nombre_unidad
        )

        return redirect('view_unidades')

    return render(request, 'unidades.html')

# Método para editar una unidad existente (requiere permisos de administrador)
@admin_required
def edit_unidad(request, unidad_id):
    unidad = get_object_or_404(Unidad, id=unidad_id)

    data = {
        'form': UnidadForm(instance=unidad)
    }

    if request.method == 'POST':
        formulario = UnidadForm(data=request.POST, instance=unidad)
        if formulario.is_valid():
            formulario.save()
            return redirect('view_unidades')
        data['form'] = formulario
    
    return render(request, 'editUnidad.html', data)

# Método para eliminar una unidad existente (requiere permisos de administrador)
@admin_required
def delete_unidad(request, unidad_id):
    unidades = get_object_or_404(Unidad, id=unidad_id)

    if request.method == 'POST':
        unidades.delete()
        return redirect('view_unidades')

    return render(request, 'unidades.html', {'unidades': unidades})

# Método para editar un producto existente (requiere permisos de administrador)
@admin_required
def edit_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    data = {
        'form': ProductoForm(instance=producto)
    }

    if request.method == 'POST':
        formulario = ProductoForm(data=request.POST, instance=producto)
        if formulario.is_valid():
            formulario.save()
            return redirect('view_productos')
        data['form'] = formulario
    
    return render(request, 'editProducto.html', data)

# Método para eliminar un producto existente (requiere permisos de administrador)
@admin_required
def delete_producto(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)

    if request.method == 'POST':
        producto.delete()
        return redirect('view_productos')

    return render(request, 'bodega.html', {'producto': producto})

# Método para gestionar una solicitud (requiere permisos de administrador)
@admin_required
def gestionar_solicitud(request, solicitud_id):
    if not request.user.is_staff:
        return redirect('view_productos')
    
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    if request.method == 'POST':
        if 'aceptar' in request.POST:
            # Aceptar la solicitud y actualizar el historial y el inventario
            solicitud.aceptada = True
            solicitud.estado = 'aceptada'
            solicitud.save()

            user_profile, created = UserProfile.objects.get_or_create(user=solicitud.usuario)
            historial = Historial.objects.create(fecha=datetime.now(), receptor=solicitud.usuario)
            for detalle in solicitud.detalles_solicitud.all():
                DetallesHistorial.objects.create(
                    historial=historial,
                    producto=detalle.producto,
                    nombre_producto=detalle.producto.nombre_producto,
                    cantidad=detalle.cantidad,
                    precio_unitario=detalle.producto.precio,
                    unidad=detalle.solicitud.usuario
                )

                user_profile.gasto_acumulado += detalle.cantidad * detalle.producto.precio
                user_profile.save()

                detalle.producto.cantidad -= detalle.cantidad
                detalle.producto.save()

            user_profile.registrar_gasto(solicitud.detalles_solicitud.all())
            solicitud.delete()
            return redirect('view_solicitudes')
        
    return redirect('view_solicitudes')

# Método para modificar la cantidad de un producto en una solicitud (requiere permisos de administrador)
def modificar_cantidad(request, solicitud_id, producto_id):
    detalle_solicitud = get_object_or_404(DetalleSolicitud, solicitud_id=solicitud_id, producto_id=producto_id)

    if request.method == 'POST':
        nueva_cantidad = int(request.POST.get('nueva_cantidad', 0))
        detalle_solicitud.cantidad = nueva_cantidad
        detalle_solicitud.save()

    return redirect('view_solicitudes')

# Método para eliminar una solicitud (requiere permisos de administrador)
@admin_required
def delete_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)

    if request.method == 'POST':
        solicitud.delete()
        return redirect('view_solicitudes')

    return render(request, 'solicitudes.html', {'solicitud': solicitud})

# Método para obtener los productos en formato de archivo Excel
def obtener_producto_excel(request):
    productos = Producto.objects.all()

    # Crear un nuevo archivo Excel y agregar los productos como filas
    workbook = Workbook()
    worksheet = workbook.active
    encabezados = ['Nombre', 'Descripcion', 'Precio', 'Cantidad', 'Categoria', 'Proveedores']
    worksheet.append(encabezados)

    for producto in productos:
        proveedores = ', '.join([proveedor.nombre_Unidad for proveedor in producto.proveedores.all()])
        fila = [producto.nombre_producto, producto.descripcion, producto.precio, producto.cantidad, producto.categoria.nombre_categoria, proveedores]
        worksheet.append(fila)

    # Ajustar el ancho de las columnas según el contenido
    for columna in worksheet.columns:
        longitud_maxima = max(len(str(celda.value)) for celda in columna)
        ajuste_ancho = (longitud_maxima + 2)
        worksheet.column_dimensions[columna[0].column_letter].width = ajuste_ancho

    # Guardar el archivo Excel y devolverlo como respuesta para descargar
    ruta_excel = 'productos.xlsx'
    workbook.save(ruta_excel)

    with open(ruta_excel, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    return response

# Método para descargar el historial de transacciones en formato de archivo Excel
def descargar_historial(request):
    historiales = Historial.objects.all()

    # Crear un nuevo archivo Excel y agregar los detalles del historial como filas
    wb = Workbook()
    ws = wb.active
    ws.append(['Fecha', 'Nombre Producto', 'Cantidad', 'Precio Unitario', 'Unidad', 'Tipo'])

    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY HH:MM:SS')

    for historial in historiales:
        detalles = DetallesHistorial.objects.filter(historial=historial)
        for detalle in detalles:
            receptor = historial.receptor.username if historial.receptor else 'N/A'
            tipo = 'Entrega' if detalle.unidad == receptor else 'Entrada'
            fecha_str = historial.fecha.strftime('%d/%m/%Y %H:%M:%S')
            ws.append([fecha_str, detalle.nombre_producto, detalle.cantidad, detalle.precio_unitario, detalle.unidad, tipo])
    
    # Ajustar el ancho de las columnas según el contenido
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Aplicar estilo de fecha a la columna de fechas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.style = date_style
            

    # Guardar el archivo Excel y devolverlo como respuesta para descargar
    ruta_excel = 'historial.xlsx'
    wb.save(ruta_excel)

    with open(ruta_excel, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=historial.xlsx'

    return response

# Método para enviar correo electrónico de notificación sobre nuevas solicitudes
def enviar_correo(solicitudes):
    subject = 'Nueva solicitud de producto(s)'
    message = 'Se han recibido nuevas solicitudes para los siguientes productos:\n'

    for solicitud in solicitudes:
        producto = solicitud.producto.nombre_producto
        cantidad = solicitud.cantidad
        message += f'- Producto: {producto}, Cantidad: {cantidad}\n'
    from_email = settings.DEFAULT_FROM_EMAIL
    admin_email = ['noresponder640@gmail.com']

    # Enviar correo electrónico
    send_mail(subject, message, from_email, admin_email, fail_silently=False)
